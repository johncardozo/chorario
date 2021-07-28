from openpyxl import Workbook
from openpyxl.worksheet.dimensions import Dimension
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from pathlib import Path
from .globales import (
    obtener_numero_franja,
    obtener_numero_dia,
    generar_franjas_descuadradas,
    color_es_oscuro,
    colores,
    franjas
)


def generar_excel_cursos(cursos):
    '''
    Genera el archivo de Excel con la asignación de profesores a cursos
    '''
    # Verifica que el directorio de salida exista, de lo contrario lo crea
    Path("output").mkdir(parents=True, exist_ok=True)

    # Crea el libro de Excel
    wb = Workbook()
    # Obtiene la hoja activa
    ws = wb.active
    # Modifica el titulo de la hoja
    ws.title = "asignacion"

    # Escribe el encabezado de la tabla
    ws.cell(row=1, column=1, value="curso")
    ws.cell(row=1, column=2, value="dia")
    ws.cell(row=1, column=3, value="hi")
    ws.cell(row=1, column=4, value="hf")
    ws.cell(row=1, column=5, value="profesor")

    fila = 2
    # Recorre los cursos
    for curso in cursos:
        # Recorre las franjas del curso
        for franja in curso["franjas"]:
            # Escribe la franja en el archivo
            ws.cell(row=fila, column=1, value=curso["curso"])
            ws.cell(row=fila, column=2, value=franja["dia"])
            ws.cell(row=fila, column=3, value=franja["hi"])
            ws.cell(row=fila, column=4, value=franja["hf"])
            # Verifica si se asignó un profesor a la franja
            if "profesor" in franja:
                ws.cell(row=fila, column=5, value=franja["profesor"])
            # Incrementa la fila del archivo
            fila = fila + 1

    # Guarda el Excel en disco
    wb.template = False
    wb.save('output/resultado_cursos.xlsx')


def generar_excel_profes(profes):
    '''
    Genera la asignación de cursos a la disponibilidad de los profesores
    '''
    # Verifica que el directorio de salida exista, de lo contrario lo crea
    Path("output").mkdir(parents=True, exist_ok=True)

    # Crea el libro de Excel
    wb = Workbook()
    # Obtiene la hoja activa
    ws = wb.active
    # Modifica el titulo de la hoja
    ws.title = "asignacion"

    # Escribe el encabezado de la tabla
    ws.cell(row=1, column=1, value="profesor")
    ws.cell(row=1, column=2, value="dia")
    ws.cell(row=1, column=3, value="hi")
    ws.cell(row=1, column=4, value="hf")
    ws.cell(row=1, column=5, value="curso")

    fila = 2
    # Recorre los profesores
    for profe in profes:
        # Recorre los dias del horario
        for dia in profe['horario']:
            # Recorre la disponibilidad del dia
            for disponibilidad in dia:
                # Escribe la disponibilidad en el archivo
                if disponibilidad['asignado']:
                    ws.cell(row=fila, column=5, value=disponibilidad["curso"])
                if disponibilidad['disponible']:
                    ws.cell(row=fila, column=1, value=profe["profesor"])
                    ws.cell(row=fila, column=2, value=disponibilidad["dia"])
                    ws.cell(row=fila, column=3, value=disponibilidad["hi"])
                    ws.cell(row=fila, column=4, value=disponibilidad["hf"])
                    #  Incrementa la fila del archivo
                    fila = fila + 1

    # Guarda el Excel en disco
    wb.template = False
    wb.save('output/resultado_profesores.xlsx')


def generar_excel_horarios(profes):
    '''
    Genera los horarios de los profesores
    '''
    # Verifica que el directorio de salida exista, de lo contrario lo crea
    Path("output").mkdir(parents=True, exist_ok=True)

    # Crea el libro de Excel
    wb = Workbook()

    # Establece el background de las celdas
    asignado_fill = PatternFill(start_color='0099CCFF',
                                end_color='0099CCFF',
                                fill_type='solid')
    no_asignado_fill = PatternFill(start_color='00FF99CC',
                                   end_color='00FF99CC',
                                   fill_type='solid')
    encabezado_fill = PatternFill(start_color='00C0C0C0',
                                  end_color='00C0C0C0',
                                  fill_type='solid')

    for profe in profes:
        # Crea una nueva hoja
        ws = wb.create_sheet(str(profe["profesor"]).lstrip().rstrip())

        # Escribe el encabezado del horario
        ws.cell(row=1, column=1, value="HI").fill = encabezado_fill
        ws.cell(row=1, column=2, value="HF").fill = encabezado_fill
        ws.cell(row=1, column=3, value="LUNES").fill = encabezado_fill
        ws.cell(row=1, column=4, value="MARTES").fill = encabezado_fill
        ws.cell(row=1, column=5, value="MIERCOLES").fill = encabezado_fill
        ws.cell(row=1, column=6, value="JUEVES").fill = encabezado_fill
        ws.cell(row=1, column=7, value="VIERNES").fill = encabezado_fill
        ws.cell(row=1, column=8, value="SABADO").fill = encabezado_fill

        # Recorre el horario del profesor
        for horario_dia in profe['horario']:
            # Recorre el dia
            for index, franja in enumerate(horario_dia):
                # Obtiene el numero de franja
                fila = obtener_numero_franja(franja['hi']) + 2

                # Escribe la hora inicio y hora fin
                ws.cell(
                    row=fila,
                    column=1,
                    value=franja['hi'])
                ws.cell(
                    row=fila,
                    column=2,
                    value=franja['hf'])

                # Obtiene el numero de columna dado el dia
                columna = obtener_numero_dia(franja['dia']) + 3

                # Escribe el valor del horario
                if franja['disponible'] and not franja['asignado']:
                    ws.cell(
                        row=fila,
                        column=columna).fill = no_asignado_fill
                if franja['disponible'] and franja['asignado']:
                    ws.cell(
                        row=fila,
                        column=columna,
                        value=franja['curso']).fill = asignado_fill

                # Elimina la hoja original
    wb.remove(wb['Sheet'])

    # Guarda el Excel en disco
    wb.template = False
    wb.save('output/resultado_horarios.xlsx')


def generar_excel_consolidado_profes(profes):
    '''
    Genera el consolidado de profesores
    '''
    # Verifica que el directorio de salida exista, de lo contrario lo crea
    Path("output").mkdir(parents=True, exist_ok=True)

    # Crea el libro de Excel
    wb = Workbook()

    # Obtiene la hoja activa
    ws = wb.active

    # Modifica el titulo de la hoja
    ws.title = "consolidado"

    # Escribe el encabezado
    ws.cell(row=1, column=1, value="PROFESOR")
    ws.cell(row=1, column=2, value="TIPO")
    ws.cell(row=1, column=3, value="CURSOS")
    ws.cell(row=1, column=4, value="CURSO 1")
    ws.cell(row=1, column=5, value="CURSO 2")
    ws.cell(row=1, column=6, value="CURSO 3")
    ws.cell(row=1, column=7, value="CURSO 4")
    ws.cell(row=1, column=8, value="CURSO 5")
    ws.cell(row=1, column=9, value="CURSO 6")

    # Recorre los profesores
    fila = 2
    for profe in profes:
        # Escribe los datos del profesor
        ws.cell(row=fila, column=1, value=profe["profesor"])
        ws.cell(row=fila, column=2, value=profe["tipo"])
        ws.cell(row=fila, column=3, value=len(profe["cursos"]))

        # Recorre los cursos
        for index, curso in enumerate(profe["cursos"]):
            ws.cell(row=fila, column=index+4, value=curso)

        # Incrementa la fila
        fila = fila + 1

    # Guarda el Excel en disco
    wb.template = False
    wb.save('output/consolidado_profesores.xlsx')


def generar_excel_cursos_sin_asignar(cursos):
    '''
    Genera un Excel con el horario de todos los cursos que no han sido asignados

    '''

    # Verifica que el directorio de salida exista, de lo contrario lo crea
    Path("output").mkdir(parents=True, exist_ok=True)

    # Crea el libro de Excel
    wb = Workbook()
    # Obtiene la hoja activa
    ws = wb.active

    # Estilo de celdas
    encabezado_fill = PatternFill(start_color='00C0C0C0',
                                  end_color='00C0C0C0',
                                  fill_type='solid')
    alineacion_central = Alignment(horizontal='center', vertical='center')
    borde = Border(left=Side(style='thin', color='000000'))

    cursos_sin_asignar = []
    franjas_por_dia = [0, 0, 0, 0, 0, 0]
    # Obtiene la cantidad de columnas por cada dia
    for curso in cursos:
        if not curso['asignado']:
            # Agrega el curso a la lista de no asignados
            cursos_sin_asignar.append(curso)
            for franja in curso['franjas']:
                # Obtiene el numero del dia
                numero_dia = obtener_numero_dia(franja['dia'])
                # Incrementa
                franjas_por_dia[numero_dia] = franjas_por_dia[numero_dia] + 1

    # Modifica el titulo de la hoja
    ws.title = f'NRC sin asignar = {len(cursos_sin_asignar)}'

    # Establece las columnas iniciales
    columnas = [
        3,  # lunes
        2 + franjas_por_dia[0] + 1,  # martes
        2 + franjas_por_dia[0] + franjas_por_dia[1] + 1,  # miercoles
        2 + franjas_por_dia[0] + franjas_por_dia[1] + \
            franjas_por_dia[2] + 1,  # jueves
        2 + franjas_por_dia[0] + franjas_por_dia[1] + \
            franjas_por_dia[2] + franjas_por_dia[3] + 1,  # viernes
        2 + franjas_por_dia[0] + franjas_por_dia[1] + franjas_por_dia[2] + \
            franjas_por_dia[3] + franjas_por_dia[4] + 1  # sabado
    ]

    # Genera una copia de las columnas originales
    columnas_originales = columnas.copy()

    # Escribe el encabezado del horario
    ws.cell(row=1, column=1, value="HI").fill = encabezado_fill
    ws.cell(row=1, column=1).alignment = alineacion_central
    ws.cell(row=1, column=2, value="HF").fill = encabezado_fill
    ws.cell(row=1, column=2).alignment = alineacion_central

    ws.cell(row=1, column=columnas[0], value="LUNES").fill = encabezado_fill
    ws.cell(row=1, column=columnas[0]).alignment = alineacion_central
    ws.merge_cells(
        start_row=1, start_column=columnas[0], end_row=1, end_column=columnas[1]-1)

    ws.cell(row=1, column=columnas[1], value="MARTES").fill = encabezado_fill
    ws.cell(row=1, column=columnas[1]).alignment = alineacion_central
    ws.merge_cells(
        start_row=1, start_column=columnas[1], end_row=1, end_column=columnas[2]-1)

    ws.cell(row=1, column=columnas[2],
            value="MIERCOLES").fill = encabezado_fill
    ws.cell(row=1, column=columnas[2]).alignment = alineacion_central
    ws.merge_cells(
        start_row=1, start_column=columnas[2], end_row=1, end_column=columnas[3]-1)

    ws.cell(row=1, column=columnas[3], value="JUEVES").fill = encabezado_fill
    ws.cell(row=1, column=columnas[3]).alignment = alineacion_central
    ws.merge_cells(
        start_row=1, start_column=columnas[3], end_row=1, end_column=columnas[4]-1)

    ws.cell(row=1, column=columnas[4], value="VIERNES").fill = encabezado_fill
    ws.cell(row=1, column=columnas[4]).alignment = alineacion_central
    ws.merge_cells(
        start_row=1, start_column=columnas[4], end_row=1, end_column=columnas[5]-1)

    ws.cell(row=1, column=columnas[5], value="SABADO").fill = encabezado_fill
    ws.cell(row=1, column=columnas[5]).alignment = alineacion_central

    # Recorre los cursos sin asignar
    index_curso = 0
    for curso in cursos_sin_asignar:
        # Crea el color de fondo de la celda
        color_fondo = colores[index_curso]

        # Crea el background de la celda
        asignado_fill = PatternFill(start_color=color_fondo,
                                    end_color=color_fondo,
                                    fill_type='solid')

        # Incrementa el numero de curso
        index_curso = index_curso + 1

        # Si el color de fondo es oscuro la fuente es clase y visceversa
        fuente = None
        if color_es_oscuro(color_fondo):
            fuente = Font(color="FFFFFF")
        else:
            fuente = Font(color="000000")

        # Recorre las franjas del curso
        for franja in curso['franjas']:
            # Obtiene las franjas normalizadas (15 minutos)
            franjas_reales = generar_franjas_descuadradas(franja)
            # Recorre las franjas normalizadas
            for franja_curso in franjas_reales:
                # Obtiene el numero de franja
                numero_franja = obtener_numero_franja(franja_curso['hi'])
                # Obtiene el numero de dia
                numero_dia = obtener_numero_dia(franja['dia'])
                # Escribe el curso en el horario
                ws.cell(row=numero_franja+2, column=1,
                        value=franja_curso['hi'])
                ws.cell(row=numero_franja+2, column=2,
                        value=franja_curso['hf'])
                # Escribe el valor de la celda y el color de fondo
                ws.cell(row=numero_franja+2,
                        column=columnas[numero_dia], value=curso['curso']).fill = asignado_fill
                # Establece el color de la fuente
                ws.cell(row=numero_franja+2,
                        column=columnas[numero_dia]).font = fuente

            # Incremente la columna del dia
            columnas[numero_dia] = columnas[numero_dia] + 1

    # Escribe las franjas en el horario
    for index, franja in enumerate(franjas):
        ws.cell(row=index+2, column=1,
                value=franja['hi']).fill = encabezado_fill
        ws.cell(row=index+2, column=1).alignment = alineacion_central
        ws.cell(row=index+2, column=2,
                value=franja['hf']).fill = encabezado_fill
        ws.cell(row=index+2, column=2).alignment = alineacion_central

    # Establece el borde entre cada uno de los días
    for columna in columnas_originales:
        for index_franja, franja in enumerate(franjas):
            ws.cell(row=index_franja+1, column=columna).border = borde
            ws.cell(row=index_franja+2, column=columna).border = borde

    # Guarda el Excel en disco
    wb.template = False
    wb.save('output/horario_cursos_sin_asignar.xlsx')


def generar_excel(cursos, profes):
    '''
    Genera la asignación resultante en archivos de Excel
    '''
    generar_excel_cursos(cursos)
    generar_excel_profes(profes)
    generar_excel_horarios(profes)
    generar_excel_consolidado_profes(profes)
    generar_excel_cursos_sin_asignar(cursos)

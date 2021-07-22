from openpyxl import load_workbook
from .globales import obtener_horario_inicial, dias

profes = []


def cargar_disponibilidad():
    '''
    Carga la disponibilidad de los profesores en el Excel. 
    return Lista de profesores de tipo diccionario.
    '''
    # Abre el libro
    wb = load_workbook(filename='input/disponibilidad.xlsx')
    # Obtiene la página activa
    ws = wb.active

    # Inicializacion de variables
    fila = 2
    # Recorre la disponibilidad de los profesores
    while ws.cell(row=fila, column=1).value is not None:
        # Crea la disponibilidad horaria del profesor
        disponibilidad = {
            "profesor": ws.cell(row=fila, column=1).value,
            "dia": ws.cell(row=fila, column=2).value,
            "hi": ws.cell(row=fila, column=3).value,
            "hf": ws.cell(row=fila, column=4).value,
            "tipo": ws.cell(row=fila, column=5).value,
        }
        # Se ubica en la siguiente fila
        fila = fila + 1
        # Agrega la disponibilidad
        agregar_disponibilidad(disponibilidad)

    return profes


def agregar_disponibilidad(disponibilidad):
    '''
    Agrega la disponibilidad
    '''
    # Busca el profesor en la lista de profesores
    profe_encontrado = next(
        (p for p in profes if p["profesor"] == disponibilidad["profesor"]), None)

    # Obtiene el numero de dia de la disponibilidad
    numero_dia = dias.index(disponibilidad['dia'])

    # Verifica si existe el profesor
    if profe_encontrado is None:
        # Crea un nuevo profesor
        nuevo_profe = {
            'profesor': disponibilidad['profesor'],
            'tipo': disponibilidad["tipo"],
            'horario': obtener_horario_inicial()
        }

        # Busca la franja del profesor para asignar la disponibilidad
        franja_disponibilidad = next(
            (f for f in nuevo_profe['horario'][numero_dia] if f['hi'] == disponibilidad['hi']), None)

        # Modifica la disponibilidad
        if franja_disponibilidad:
            franja_disponibilidad['disponible'] = True
            #franja_disponibilidad['profesor'] = disponibilidad['profesor']

        # Agrega el profesor a la lista
        profes.append(nuevo_profe)

    else:
        # Busca la franja del profesor para asignar la disponibilidad
        franja_disponibilidad = next(
            (f for f in profe_encontrado['horario'][numero_dia] if f['hi'] == disponibilidad['hi']), None)

        # Modifica la disponibilidad
        if franja_disponibilidad:
            franja_disponibilidad['disponible'] = True
            #franja_disponibilidad['profesor'] = disponibilidad['profesor']

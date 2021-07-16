from openpyxl import load_workbook
chorario = []

def cargar_cursos():
    '''
    Carga los cursos del chorario que se encuentran en Excel. 
    return Lista de cursos de tipo diccionario.
    '''
    # Abre el libro
    wb = load_workbook(filename = 'input/cursos.xlsx')
    # Obtiene la página activa
    ws = wb.active

    # Inicializacion de variables
    fila = 2
    # Recorre los cursos
    while ws.cell(row=fila, column=1).value is not None:
        # Crea la franja horaria del curso
        franja = {
            "curso": ws.cell(row=fila, column=1).value,
            "dia" : ws.cell(row=fila, column=2).value,
            "hi" : ws.cell(row=fila, column=3).value,
            "hf" : ws.cell(row=fila, column=4).value
        }
        # Se ubica en la siguiente fila
        fila = fila + 1
        # Agrega la franja al chorario
        agregar_franja(franja)
    
    return chorario


def agregar_franja(franja):
    '''
    Agrega la franja de un curso al chorario
    '''
    # Intenta encontrar el curso en el chorario
    curso_encontrado = next((c for c in chorario if c["curso"] == franja["curso"]), None)

    # Crea la nueva franja
    nueva_franja = {
        "dia": franja["dia"],
        "hi": franja["hi"],
        "hf": franja["hf"],
        "asignada": False
    }

    # Verifica si existe el curso en el chorario
    if curso_encontrado is None:
        # Crea el nuevo curso
        nuevo_curso = {
            "curso": franja["curso"],
            "franjas": [ nueva_franja ]
        }
        # Agrega el curso al chorario
        chorario.append(nuevo_curso)
    else:
        # Agrega la disponibilidad al curso encontrado en el chorario
        curso_encontrado["franjas"].append(nueva_franja)

    